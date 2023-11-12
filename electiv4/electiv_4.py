import random
import time
import openpyxl


def selection_sort(numbers: list[int]) -> list[int]:
    for i in range(len(numbers) - 1):
        min_index = i
        for j in range(i + 1, len(numbers)):
            if numbers[j] < numbers[min_index]:
                min_index = j

        if min_index != i:
            numbers[min_index], numbers[i] = numbers[i], numbers[min_index]
    return numbers


def insertion_sort(numbers: list[int]) -> list[int]:
    for i in range(1, len(numbers)):
        current_value = numbers[i]
        j = i - 1
        while j >= 0:
            if numbers[j] > current_value:
                numbers[j + 1] = numbers[j]
                numbers[j] = current_value
                j -= 1
            else:
                break

    return numbers


def bubble_sort(numbers: list[int]) -> list[int]:
    for i in range(len(numbers)):
        for j in range(len(numbers) - i - 1):
            if numbers[j] > numbers[j + 1]:
                numbers[j], numbers[j + 1] = numbers[j + 1], numbers[j]
    return numbers


def shell_sort(numbers: list[int]) -> list[int]:
    gap = len(numbers) // 2

    while gap > 0:
        for i in range(gap, len(numbers)):
            current_value = numbers[i]
            position = i

            while position >= gap and numbers[position - gap] > current_value:
                numbers[position], numbers[position - gap] = numbers[position - gap], numbers[position]
                position -= gap

        gap //= 2
    return numbers


def quick_sort(numbers: list[int]) -> list[int]:
    n = len(numbers)
    if n <= 1:
        return numbers
    else:
        pivot = numbers[len(numbers) // 2]
        less = [x for x in numbers if x < pivot]
        greater_or_equal = [x for x in numbers if x > pivot]
        return quick_sort(less) + [pivot] + quick_sort(greater_or_equal)


numbers = [random.randint(-20, 20) for i in range(20)]

sort_functions = [selection_sort, insertion_sort, bubble_sort, shell_sort, quick_sort]
sort_names = ['Сортировка выбором', 'Сортировка вставками', 'Пузырьковая сортировка',
              'Сортировка Шелла', 'Быстрая сортировка']

wb = openpyxl.Workbook()
ws = wb.active
ws.cell(row=1, column=1).value = "Неотсортированный массив"
for i in range(len(numbers)):
    ws.cell(row=i+2, column=1).value = numbers[i]


for j in range(len(sort_functions)):
    copy = numbers[:]
    start = time.perf_counter()
    sort_functions[j](copy)
    end = time.perf_counter()
    duration = end - start
    ws.cell(row=1, column=j+2).value = f"{sort_names[j]} {duration} сек"
    for i in range(len(copy)):
        ws.cell(row=i+2, column=j+2).value = copy[i]

wb.save("sorted_arrays.xlsx")
